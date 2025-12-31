"""Tests for Excel parser edge cases."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from static_traffic_analyzer.parsers.excel import parse_excel


def test_excel_member_empty_lines(tmp_path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    address_sheet = workbook.create_sheet("Address Object")
    address_sheet.append(["Object Name", "Type", "Subnet/Start-IP", "Mask/End-IP"])
    address_sheet.append(["net", "ipmask", "10.0.0.0", "255.255.255.0"])

    address_group_sheet = workbook.create_sheet("Address Group")
    address_group_sheet.append(["Group Name", "Member"])
    address_group_sheet.append(["group", "net\n\n"])

    service_group_sheet = workbook.create_sheet("Service Group")
    service_group_sheet.append(["Group Name", "Member"])
    service_group_sheet.append(["svc-group", "tcp_80\n\n"])

    rule_sheet = workbook.create_sheet("Rule")
    rule_sheet.append(["Seq", "Enable", "Source", "Destination", "Service", "Action", "ID", "Comments"])
    rule_sheet.append([1, True, "group", "group", "svc-group", "accept", "1", "test"])

    path = tmp_path / "rules.xlsx"
    workbook.save(path)

    data = parse_excel(str(path))

    assert data.address_book.groups["group"].members == ("net",)
    assert data.service_book.groups["svc-group"].members == ("tcp_80",)
    assert data.policies[0].services == ("svc-group",)
